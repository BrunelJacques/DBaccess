#!/usr/bin/env python
# -*- coding: utf-8 -*-
# ------------------------------------------------------------------------
# Application :    Noethys, gestion multi-activités
# Site internet :  www.noethys.com
# Auteur:           Ivan LUCAS
# Copyright:       (c) 2010-11 Ivan LUCAS
# Licence:         Licence GNU GPL
# ------------------------------------------------------------------------

import wx
import os
import datetime
import decimal
import platform

from xpy.outils import xchemins, xformat


class DataType(object):
    #Classe permetant la conversion facile vers le format souhaité (nombre de caractéres, alignement, décimales)
    def __init__(self,cat=int,lg=1,align="<",precision=2,fmt=None,**kwd):
        """
        initialise l'objet avec les paramétres souhaité
        """
        self.cat = cat
        self.length = lg
        self.align = align
        self.precision = precision
        self.fmt = fmt
        if self.cat == 'const':
            self.constante = kwd.pop('constante',None)

    def Convert(self,data):
        # convertit au format souhaité
        ret_val = ""

        # gestion des valeurs nulles selon la catégorie attendue
        if data == None or data == '':
            if self.cat in (int,float,bool,decimal.Decimal):
                data = self.cat(0)
            elif self.cat == str:
                data = ''

        # avec un format spécifique fourni
        if self.fmt: # attention la catégorie doit bien correspondre à la réalité attendue par le format
            try:
                if self.cat == wx.DateTime:
                    # exemple data.Format("%d%m%y")
                    ret_val = data.Format(self.fmt)
                elif self.cat == datetime.date:
                    # exemple '{:%d%m%y}'.format(data)
                    ret_val = self.fmt.format(data)
                else:
                    # exemple "{0:{align}0{length}.{precision}f}".format(+1254.126,align=">",length=10,precision=2)
                    # ou directement "{0:+010.0f}".format(+1254.126) pour '+000001254'
                    ret_val = self.fmt.format(data,align=self.align,length=self.length,prec=self.precision)
            except:
                    ret_val = ' ' * self.length

        elif self.cat == int:                        #si l'on veux des entier
            if data!="":
                try:                                #on vérifie qu'il s'agit bien d'un nombre
                    data=int(data)
                except ValueError as e:
                    print("/!\\ Erreur de format, impossible de convertir en int /!\\")
                    print(e)
                    data=0
                ret_val = u"{0:{align}0{length}d}".format(data,align=self.align,length=self.length)
            else:
                ret_val = u"{0:{align}0{length}s}".format(data,align=self.align,length=self.length)

        elif self.cat == str:                      #si l'on veux des chaines de caractéres
            if not isinstance(data,(str)): data = str(data)
            for a in ['\\',';',',']:
                data = data.replace(a,'')
            ret_val = u"{0: {align}0{length}s}".format(data,align=self.align,length=self.length)

        elif self.cat == 'const':                      # la valeur est une constante
            if not isinstance(data,(str)): data = str(data)
            ret_val = self.constante

        elif self.cat == float:                    #si l'on veux un nombre a virgule
            if data!="":
                try:
                    data=float(data)
                    #on vérifie qu'il s'agit bien d'un nombre
                except ValueError as e:
                    print("/!\\ Erreur de format, impossible de convertir en float /!\\")
                    print(e)
                    data=0
                ret_val = u"{0: {align}0{length}.{precision}f}".format(data,align=self.align,length=self.length,precision=self.precision)
            else:
                ret_val = u"{0: {align}0{length}s}".format(data,align=self.align,length=self.length)

        if len(ret_val)>self.length:                #on tronc si la chaine est trop longue
            ret_val=ret_val[:self.length]
        return ret_val
        #fin Convert
    #fin class DataType

def GetValeursListview(listview=None, format="texte"):
    """ Récupère les valeurs affichées sous forme de liste """
    """ format = "texte" ou "original" """
    # Récupère les labels de colonnes
    lstColonnes = []
    ixID = None
    for colonne in listview.columns:
        lstColonnes.append((colonne.title, colonne.align, colonne.width, colonne.valueGetter))
        if not ixID and isinstance(colonne.valueGetter,str) and colonne.valueGetter.startswith('ID'):
            ixID = listview.columns.index(colonne)

    # Récupère les valeurs
    listeValeurs = []
    listeObjects = []
    if listview.checkColonne:
        listeObjects = listview.GetCheckedObjects()
    if len(listeObjects) == 0:
        listeObjects = listview.GetSelectedObjects()
    if len(listeObjects) <= 1:
        listeObjects = listview.innerList  # listview.GetFilteredObjects()
    for object in listeObjects:
        # élude les lignes sans valeur sur la 1ere colonne ID
        if ixID:
            ID = listview.GetStringValueAt(object, ixID)
            if not ID or ID == 0:
                continue
        valeursLigne = []
        for indexCol in range(0, listview.GetColumnCount()):
            if format == "texte":
                valeur = listview.GetStringValueAt(object, indexCol)
            else:
                valeur = listview.GetValueAt(object, indexCol)
            valeursLigne.append(valeur)
        listeValeurs.append(valeursLigne)

    return lstColonnes, listeValeurs

def GetValeursGrid(grid=None):
    """ Récupère les valeurs affichées sous forme de liste """
    # Récupère les labels de colonnes
    lstColonnes = [("titre_ligne", None, grid.GetColLabelSize(), "titre_ligne"), ]
    for numCol in range(0, grid.GetNumberCols()):
        lstColonnes.append(
            (grid.GetColLabelValue(numCol), None, grid.GetColSize(numCol), grid.GetColLabelValue(numCol)))

    # Récupère les valeurs
    listeValeurs = []
    for numLigne in range(0, grid.GetNumberRows()):
        labelLigne = grid.GetRowLabelValue(numLigne)
        valeursLigne = [labelLigne, ]
        for numCol in range(0, grid.GetNumberCols()):
            valeur = grid.GetCellValue(numLigne, numCol)
            if type(valeur) not in ("str", "unicode"):
                valeur = str(valeur)
            valeursLigne.append(valeur)
        listeValeurs.append(valeursLigne)

    return lstColonnes, listeValeurs

def ChoixDestination(nomFichier,wildcard):
    # Demande à l'utilisateur le nom de fichier et le répertoire de destination
    sp = wx.StandardPaths.Get()
    cheminDefaut = sp.GetDocumentsDir()
    dlg = wx.FileDialog(
        None, message="Veuillez sélectionner le répertoire de destination et le nom du fichier",
        defaultDir=cheminDefaut,
        defaultFile=nomFichier,
        wildcard=wildcard,
        style=wx.FD_SAVE
    )
    dlg.SetFilterIndex(0)
    if dlg.ShowModal() == wx.ID_OK:
        cheminFichier = dlg.GetPath()
        dlg.Destroy()
    else:
        dlg.Destroy()
        return

    # Le fichier de destination existe déjà :
    if os.path.isfile(cheminFichier) == True:
        dlg = wx.MessageDialog(None, "Un fichier portant ce nom existe déjà. \n\nVoulez-vous le remplacer ?",
                               "Attention !", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_EXCLAMATION)
        if dlg.ShowModal() == wx.ID_NO:
            dlg.Destroy()
            return False
        else:
            dlg.Destroy()
    return cheminFichier

def Confirmation(cheminFichier):
    # Confirmation de création du fichier et demande d'ouverture directe
    txtMessage = "Le fichier a été créé avec succès. Souhaitez-vous l'ouvrir dès maintenant ?"
    dlgConfirm = wx.MessageDialog(None, txtMessage, "Confirmation", wx.YES_NO | wx.NO_DEFAULT | wx.ICON_QUESTION)
    reponse = dlgConfirm.ShowModal()
    dlgConfirm.Destroy()
    if reponse == wx.ID_YES:
        LanceFichierExterne(cheminFichier)
    return wx.OK

def LigneLgFixe(matrice):
    # crée une fonction appelant les datatype pour formater une ligne à données fixes
    # la matrice décrivant les params doit être structurée: {'code': 'champ1', 'typ': str, 'lg': 8, 'align': "<"},
    lstFunc = [DataType(**x).Convert for x in matrice]
    lstChamps = [x['code'] for x in matrice]

    def func(valeurs):
        texte = ''
        if len(valeurs) != len(lstFunc):
            wx.MessageBox('La matrice prévoit %d champs, la ligne a %d valeurs\n%s\n%s'%(len(lstFunc),len(valeurs),
                                                                                        str(lstChamps),
                                                                                        str(valeurs)),
                          'Echec xexport.LigneLgFixe')
        for ix in range(len(lstFunc)):
            texte += lstFunc[ix](valeurs[ix])
        return texte

    return func

# -------------------------------------------------------------------------------------------------------------------------------

def LanceFichierExterne(nomFichier) :
    """ Ouvre un fichier externe sous windows ou linux """
    if platform.system() == "Windows":
        nomFichier = nomFichier.replace("/", "\\")
        os.startfile(nomFichier)
    if platform.system() == "Linux":
        os.system("xdg-open " + nomFichier)

def ComposeTexte(lstColonnes,lstValeurs):
    # Création du fichier texte
    texte = ""
    separateur = ";"
    for labelCol, alignement, largeur, code in lstColonnes:
        try:
            if "CheckState" in str(code):
                code = "Coche"
        except:
            pass
        texte += labelCol + separateur
    texte = texte[:-1] + "\n"

    for valeurs in lstValeurs:
        for valeur in valeurs:
            if valeur == None:
                valeur = ""
            texte += "%s%s" % (valeur, separateur)
        texte = texte[:-1] + "\n"
    # Elimination du dernier saut à la ligne
    return texte[:-1]

def ExportTexte(listview=None, grid=None, titre=u"",lstColonnes=None,listeValeurs=None):
    """ Export de la liste au format texte """
    if (listview != None and len(listview.innerList) == 0) or (
            grid != None and (grid.GetNumberRows() == 0 or grid.GetNumberCols() == 0)):
        dlg = wx.MessageDialog(None, "Il n'y a aucune donnée dans la liste !", "Erreur", wx.OK | wx.ICON_ERROR)
        dlg.ShowModal()
        dlg.Destroy()
        return wx.CANCEL

    # Récupération des valeurs
    if listview != None and listeValeurs == None:
        lstColonnes, listeValeurs = GetValeursListview(listview, format="texte")

    if grid != None and lstColonnes == None and listeValeurs == None:
        lstColonnes, listeValeurs = GetValeursGrid(grid)

    nomFichier = "ExportTexte_%s.txt" % datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    wildcard = "Fichier texte (*.txt)|*.txt|" \
               "All files (*.*)|*.*"
    cheminFichier = ChoixDestination(nomFichier,wildcard)
    if not cheminFichier : return wx.CANCEL

    # Création du fichier texte
    texte = titre + "\n"
    texte += ComposeTexte(lstColonnes,listeValeurs)
    f = open(cheminFichier, "w")
    f.write(texte)
    f.close()
    return Confirmation(cheminFichier)

def ExportLgFixe(nomfic='',matrice={},valeurs=[],entete=False):
    """ Export de la liste au format texte """
    if len(valeurs) == 0:
        wx.MessageBox("On ne peut exporter une liste de valeurs vide")
        return wx.CANCEL
    if len(matrice) != len(valeurs[0]):
        wx.MessageBox("Pb: matrice décrit %d colonnes et valeurs[0] en contient %d!"%(len(matrice),
                                                                                        len(valeurs)[0]))

    # Demande à l'utilisateur le nom de fichier et le répertoire de destination
    nomFichier = "%s"%nomfic
    wildcard = "Fichier texte (*.txt)|*.txt|" \
               "All files (*.*)|*.*"
    cheminFichier = ChoixDestination(nomFichier,wildcard)
    if not cheminFichier : return wx.CANCEL

    # Création du fichier texte
    texte = ""
    if entete:
        ligne = [x['code'] for x in matrice]
        matriceEntete = [{'code':x['code'],'cat':str,'lg':x['lg'],} for x in matrice]
        texte += LigneLgFixe(matriceEntete)(ligne)+ "\n"

    makeLigne= LigneLgFixe(matrice)
    for ligne in valeurs:
        texte += makeLigne(ligne)
        texte = texte[:-1] + "\n"

    # Elimination du dernier saut à la ligne
    texte = texte[:-1]

    # Création du fichier texte
    f = open(cheminFichier, "w")
    f.write(texte)
    f.close()
    return Confirmation(cheminFichier)

def ExportExcel(listview=None, grid=None, titre="Liste", lstColonnes=None, listeValeurs=None, autoriseSelections=False):
    # Export de la liste au format Excel

    # Vérifie si données bien présentes
    if (listview != None and len(listview.innerList) == 0) or (
            grid != None and (grid.GetNumberRows() == 0 or grid.GetNumberCols() == 0)):
        dlg = wx.MessageDialog(None, "Il n'y a aucune donnée dans la liste !", "Erreur", wx.OK | wx.ICON_ERROR)
        dlg.ShowModal()
        dlg.Destroy()
        return wx.CANCEL

    # Récupération des valeurs
    if listview != None and lstColonnes == None and listeValeurs == None:
        lstColonnes, listeValeurs = GetValeursListview(listview, format="original")

    if grid != None and lstColonnes == None and listeValeurs == None:
        autoriseSelections = False
        lstColonnes, listeValeurs = GetValeursGrid(grid)

    # Définit le nom et le chemin du fichier
    nomFichier = "ExportExcel_%s.xls" % datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    wildcard = "Fichier Excel (*.xls)|*.xls|" \
               "All files (*.*)|*.*"
    cheminFichier = ChoixDestination(nomFichier,wildcard)
    if not cheminFichier : return wx.CANCEL

    # Export
    # Création d'un classeur
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, NamedStyle
    from openpyxl.utils import get_column_letter

    # Create workbook and sheet
    wb = Workbook()
    title = xformat.NoPunctuation(titre)
    ws1 = wb.active
    ws1.title = title[:25]

    # Define alignments
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # Define styles
    style_euros = NamedStyle(name="style_euros")
    style_euros.number_format = '"$"#,##0.00_);("$"#,##'
    style_euros.alignment = align_right

    style_date = NamedStyle(name="style_date")
    style_date.number_format = 'DD/MM/YYYY'
    style_date.alignment = align_right

    style_heure = NamedStyle(name="style_heure")
    style_heure.number_format = '[hh]:mm'
    style_heure.alignment = align_right

    # Register styles (optional but good practice)
    for style in [style_euros, style_date, style_heure]:
        if style.name not in wb.named_styles:
            wb.add_named_style(style)

    # Write column labels
    x = 1  # openpyxl is 1-indexed
    y = 1
    for labelCol, alignement, largeur, nomChamp in lstColonnes:
        try:
            if "CheckState" in nomChamp:
                nomChamp = "Coche"
        except:
            pass

        cell = ws1.cell(row=x, column=y, value=labelCol)

        # Apply alignment
        if alignement == 'left':
            cell.alignment = align_left
        else:
            cell.alignment = align_right

        # Set column width
        if largeur <= 0:
            largeur = 1
        col_letter = get_column_letter(y)
        ws1.column_dimensions[col_letter].width = largeur

        y += 1

    # Save workbook
    wb.save("output.xlsx")

    # -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    # Création des lignes
    def RechercheFormatFromChaine(valeur):
        """ Recherche le type de la chaîne """
        if valeur.endswith(SYMBOLE):
            # Si c'est un montant en euros
            try:
                if valeur.startswith("- "):
                    valeur = valeur.replace("- ", "-")
                if valeur.startswith("+ "):
                    valeur = valeur.replace("+ ", "")
                nbre = float(valeur[:-1])
                return (nbre, styleEuros)
            except:
                pass

        # Si c'est un nombre
        try:
            if valeur.startswith("- "):
                valeur = valeur.replace("- ", "-")
            nbre = float(valeur)
            return (nbre, None)
        except:
            pass

        # Si c'est une date
        try:
            if len(valeur) == 10:
                if valeur[2] == "/" and valeur[5] == "/":
                    return (valeur, styleDate)
        except:
            pass

        if type(valeur) == datetime.timedelta:
            return (valeur, styleHeure)

        # Si c'est une heure
        try:
            if len(valeur) > 3:
                if ":" in valeur:
                    separateur = ":"
                elif "h" in valeur:
                    separateur = "h"
                else:
                    separateur = None
                if separateur != None:
                    heures, minutes = valeur.split(separateur)
                    valeur = datetime.timedelta(minutes=int(heures) * 60 + int(minutes))
                    # valeur = datetime.time(hour=int(valeur.split(separateur)[0]), minute=int(valeur.split(separateur)[1]))
                    return (valeur, styleHeure)
        except:
            pass

        return str(valeur), None

    # -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    def RechercheFormat(valeur):
        """ Recherche le type de la donnée """
        if type(valeur) == decimal.Decimal:
            valeur = float(valeur)
            return (valeur, styleEuros)

        if type(valeur) == float:
            return (valeur, None)

        if type(valeur) == int:
            return (valeur, None)

        if type(valeur) == datetime.date:
            valeur = xformat.DatetimeToStr(valeur)
            return (valeur, styleDate)

        if type(valeur) == datetime.timedelta:
            return (valeur, styleHeure)

        try:
            if len(valeur) > 3:
                if ":" in valeur:
                    separateur = ":"
                elif "h" in valeur:
                    separateur = "h"
                else:
                    separateur = None
                if separateur != None:
                    donnees = valeur.split(separateur)
                    if len(donnees) == 2:
                        heures, minutes = donnees
                    if len(donnees) == 3:
                        heures, minutes, secondes = donnees
                    valeur = datetime.timedelta(minutes=int(heures) * 60 + int(minutes))
                    # valeur = datetime.time(hour=int(valeur.split(separateur)[0]), minute=int(valeur.split(separateur)[1]))
                    return (valeur, styleHeure)
        except:
            pass

        if type(valeur) in (str,):
            if len(valeur) == 10:
                if valeur[2] == "/" and valeur[5] == "/": return (valeur, styleDate)
                if valeur[4] == "-" and valeur[7] == "-": return (
                xformat.DateSqlToIso(valeur), styleDate)

        return str(valeur), None

    # -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------

    x = 1
    y = 0
    for valeurs in listeValeurs:
        if autoriseSelections == False or int(valeurs[0]) in listeSelections:
            for valeur in valeurs:
                if valeur == None:
                    valeur = u""

                # Recherche s'il y a un format de nombre ou de montant
                if isinstance(valeur,str):
                    valeur, format = RechercheFormatFromChaine(valeur)  # RechercheFormatFromChaine(valeur)
                else:
                    valeur, format = RechercheFormat(valeur)

                # Enregistre la valeur
                if format != None:
                    ws1.write(x, y, valeur, format)
                else:
                    ws1.write(x, y, valeur)

                y += 1
            x += 1
            y = 0

    # Finalisation du fichier xls
    try:
        wb.save(cheminFichier)
    except:
        dlg = wx.MessageDialog(None,
            "Il est impossible d'enregistrer le fichier Excel. Veuillez vérifier que ce fichier n'est pas déjà ouvert en arrière-plan.",
                               "Erreur", wx.OK | wx.ICON_ERROR)
        dlg.ShowModal()
        dlg.Destroy()
        return wx.CANCEL

    return Confirmation(cheminFichier)

def ExportTemp(lstColonnes,llData,nomFichier="spare.txt"):
    chemin = xchemins.GetRepTemp(nomFichier)
    texte = ComposeTexte(lstColonnes,llData)
    # Création du fichier texte
    f = open(chemin, "w")
    f.write(texte)
    f.close()
    return

def ImportTemp(nomFichier="spare.txt"):
    chemin = xchemins.GetRepTemp(nomFichier)
    # ouverture du fichier texte
    f = open(chemin, "r")
    ldDon = []
    for line in f:
        ldDon.append(line)
    f.close()
    return ldDon

# ------------------------- POUR LES TESTS ---------------------------------------------


if __name__ == '__main__':
    app = wx.App(0)
    os.chdir("../..")
    os.chdir("../..")
    # wx.InitAllImageHandlers()
    frame_1 = wx.Frame(None, -1, "OL Test Export")
    app.SetTopWindow(frame_1)
    frame_1.Show()
    app.MainLoop()
